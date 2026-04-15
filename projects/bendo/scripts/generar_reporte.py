"""
generar_reporte.py
Genera resumen_bendo.xlsx con 3 hojas:
  1. Resumen — descriptivos generales de base_presentar
  2. Glosario — diccionario de columnas
  3. Base Presentar — datos completos
Python 3.11 / Polars / openpyxl
"""

import polars as pl
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASES_DIR = Path(__file__).resolve().parent.parent / "bases"
ROOT_DIR = Path(__file__).resolve().parent.parent
BASE_PRESENTAR = BASES_DIR / "base_presentar_contactabilidad.parquet"
RESUMEN_XLSX = ROOT_DIR / "resumen_bendo.xlsx"

# ── Estilos ──────────────────────────────────────────────────────────
FONT_TITULO = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_BOLD = Font(name="Calibri", size=10, bold=True)
FONT_SUBTITULO = Font(name="Calibri", size=11, bold=True, color="1F4E79")

FILL_TITULO = PatternFill("solid", fgColor="1F4E79")
FILL_HEADER = PatternFill("solid", fgColor="2E75B6")
FILL_LIGHT = PatternFill("solid", fgColor="D6E4F0")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_GREEN = PatternFill("solid", fgColor="E2EFDA")
FILL_YELLOW = PatternFill("solid", fgColor="FFF2CC")
FILL_RED = PatternFill("solid", fgColor="FCE4EC")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

BORDER_THIN = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)


def apply_style(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def write_title(ws, row, col, text, colspan=1):
    cell = ws.cell(row=row, column=col, value=text)
    apply_style(cell, FONT_TITULO, FILL_TITULO, ALIGN_CENTER, BORDER_THIN)
    if colspan > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + colspan - 1,
        )
        for c in range(col + 1, col + colspan):
            apply_style(ws.cell(row=row, column=c), fill=FILL_TITULO, border=BORDER_THIN)


def write_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        apply_style(cell, FONT_HEADER, FILL_HEADER, ALIGN_CENTER, BORDER_THIN)


def write_row(ws, row, values, start_col=1, fills=None, fonts=None, fmt=None):
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=v)
        f = fills[i] if fills and i < len(fills) else FILL_WHITE
        fn = fonts[i] if fonts and i < len(fonts) else FONT_NORMAL
        al = ALIGN_RIGHT if isinstance(v, (int, float)) else ALIGN_LEFT
        nf = None
        if fmt and i < len(fmt) and fmt[i]:
            nf = fmt[i]
        apply_style(cell, fn, f, al, BORDER_THIN, nf)


def pct(part, total):
    return part / total if total > 0 else 0


def write_section_title(ws, row, text, ncols):
    cell = ws.cell(row=row, column=1, value=text)
    apply_style(cell, FONT_SUBTITULO, FILL_TITULO, ALIGN_LEFT, BORDER_THIN)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        for c in range(2, ncols + 1):
            apply_style(ws.cell(row=row, column=c), fill=FILL_TITULO, border=BORDER_THIN)


# ── Hoja 1: Resumen ─────────────────────────────────────────────────
def hoja_resumen(wb, df_pres):
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_properties.tabColor = "1F4E79"

    NCOLS = 4
    r = 1
    write_title(ws, r, 1, "RESUMEN - BASE COMERCIAL BENDO", NCOLS)

    # ── 1. Descriptivos generales ────────────────────────────────────
    n_rucs = df_pres["numero_ruc"].n_unique()
    n_tipo_act = df_pres["tipo_actividad"].n_unique()
    n_con_balance = df_pres.filter(pl.col("valor_balance_2024").is_not_null()).height

    r += 2
    write_title(ws, r, 1, "1. DESCRIPTIVOS GENERALES", NCOLS)
    r += 1
    write_headers(ws, r, ["Metrica", "Valor", "", ""])
    r += 1

    stats = [
        ("RUCs unicos", n_rucs, "#,##0"),
        ("Tipos de actividad", n_tipo_act, "#,##0"),
        ("RUCs con balance 2024", n_con_balance, "#,##0"),
    ]
    for label, val, nf in stats:
        fill = FILL_LIGHT if (r % 2 == 0) else FILL_WHITE
        write_row(ws, r, [label, val, "", ""],
                  fills=[fill] * NCOLS, fonts=[FONT_BOLD, FONT_NORMAL, FONT_NORMAL, FONT_NORMAL],
                  fmt=[None, nf, None, None])
        r += 1

    # ── 2. Composición por tipo de contribuyente ─────────────────────
    r += 1
    NCOLS_TIPO = 4
    write_title(ws, r, 1, "2. COMPOSICION POR TIPO DE CONTRIBUYENTE", NCOLS_TIPO)
    r += 1
    write_headers(ws, r, [
        "Tipo de Contribuyente", "RUCs", "% del Total", "Ingreso Prom. Estimado ($)",
    ])
    r += 1

    tipo_contrib = (
        df_pres.group_by("tipo_contribuyente")
        .agg(
            pl.col("numero_ruc").n_unique().alias("rucs"),
            pl.col("ingreso_estimado_2025").mean().alias("ingreso_prom"),
        )
        .sort("rucs", descending=True)
    )

    for row_data in tipo_contrib.iter_rows(named=True):
        fill = FILL_LIGHT if (r % 2 == 0) else FILL_WHITE
        ingreso = round(row_data["ingreso_prom"], 2) if row_data["ingreso_prom"] else 0
        write_row(ws, r, [
            row_data["tipo_contribuyente"],
            row_data["rucs"],
            pct(row_data["rucs"], n_rucs),
            ingreso,
        ],
            fills=[fill] * NCOLS_TIPO,
            fonts=[FONT_BOLD, FONT_NORMAL, FONT_NORMAL, FONT_NORMAL],
            fmt=[None, "#,##0", "0.0%", "$#,##0.00"],
        )
        r += 1

    write_row(ws, r, ["TOTAL", n_rucs, 1.0, round(df_pres["ingreso_estimado_2025"].mean(), 2)],
              fills=[FILL_HEADER] * NCOLS_TIPO, fonts=[FONT_HEADER] * NCOLS_TIPO,
              fmt=[None, "#,##0", "0.0%", "$#,##0.00"])
    r += 1

    # ── 3. Ingreso y ticket promedio por tipo_actividad ──────────────
    r += 1
    NCOLS_ACT = 5
    write_title(ws, r, 1, "3. INGRESO Y TICKET PROMEDIO POR TIPO DE ACTIVIDAD", NCOLS_ACT)
    r += 1
    write_headers(ws, r, [
        "Tipo de Actividad", "RUCs", "% del Total",
        "Ingreso Prom. Estimado ($)", "Ticket Mediana ($)",
    ])
    r += 1

    tipo_act = (
        df_pres.group_by("tipo_actividad")
        .agg(
            pl.col("numero_ruc").n_unique().alias("rucs"),
            pl.col("ingreso_estimado_2025").mean().alias("ingreso_prom"),
            pl.col("ticket_promedio_2025").median().alias("ticket_prom"),
        )
        .sort("rucs", descending=True)
    )

    for row_data in tipo_act.iter_rows(named=True):
        is_top = row_data["rucs"] >= 100
        fill = FILL_LIGHT if is_top else FILL_WHITE
        ingreso = round(row_data["ingreso_prom"], 2) if row_data["ingreso_prom"] else 0
        ticket = round(row_data["ticket_prom"], 2) if row_data["ticket_prom"] else 0
        write_row(ws, r, [
            row_data["tipo_actividad"],
            row_data["rucs"],
            pct(row_data["rucs"], n_rucs),
            ingreso,
            ticket,
        ],
            fills=[fill] * NCOLS_ACT,
            fonts=[FONT_BOLD if is_top else FONT_NORMAL] + [FONT_NORMAL] * 4,
            fmt=[None, "#,##0", "0.0%", "$#,##0.00", "$#,##0.00"],
        )
        r += 1

    # Total
    ingreso_total = df_pres["ingreso_estimado_2025"].mean()
    ticket_total = df_pres["ticket_promedio_2025"].median()
    write_row(ws, r, [
        "TOTAL", n_rucs, 1.0,
        round(ingreso_total, 2) if ingreso_total else 0,
        round(ticket_total, 2) if ticket_total else 0,
    ],
        fills=[FILL_HEADER] * NCOLS_ACT,
        fonts=[FONT_HEADER] * NCOLS_ACT,
        fmt=[None, "#,##0", "0.0%", "$#,##0.00", "$#,##0.00"],
    )

    # Ajustar anchos
    col_widths = [38, 20, 16, 28, 22]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A2"


# ── Hoja 2: Glosario ────────────────────────────────────────────────
def hoja_glosario(wb):
    ws = wb.create_sheet("Glosario")
    ws.sheet_properties.tabColor = "2E75B6"

    glosario = [
        ("numero_ruc", "int", "Registro Unico de Contribuyentes del SRI"),
        ("razon_social", "str", "Nombre legal registrado en el SRI"),
        ("nombre_comercial", "str", "Nombre comercial derivado de los establecimientos (top 3 palabras mas frecuentes)"),
        ("clase_contribuyente", "str", "Clasificacion tributaria: RIMPE, GENERAL, ESPECIAL, etc."),
        ("tipo_contribuyente", "str", "PERSONA NATURAL o SOCIEDAD"),
        ("actividad_economica", "str", "Descripcion de la actividad economica principal (CIIU)"),
        ("valor_balance_2024", "float", "Valor del balance declarado en 2024 (USD)"),
        ("ticket_promedio_2025", "float", "Promedio del valor por factura en USD (media de todos los establecimientos)"),
        ("ingreso_estimado_2025", "float",
         "Ingreso estimado en USD (suma de todos los establecimientos del RUC, "
         "con imputacion de outliers bajos y medianas por grupo)"),
        ("tipo_actividad", "str", "Categoria agrupada de actividad economica"),
        ("descripcion_corta", "str", "Descripcion resumida de la actividad economica"),
    ]

    write_title(ws, 1, 1, "GLOSARIO DE COLUMNAS - BASE PRESENTAR", 3)
    write_headers(ws, 2, ["Columna", "Tipo", "Descripcion"])

    for i, (col_name, tipo, desc) in enumerate(glosario, start=3):
        fill = FILL_LIGHT if i % 2 == 0 else FILL_WHITE
        write_row(ws, i, [col_name, tipo, desc], fills=[fill] * 3)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 90
    ws.freeze_panes = "A3"


# ── Hoja 3: Base Presentar (datos) ──────────────────────────────────
def hoja_base_presentar(wb, df_pres):
    ws = wb.create_sheet("Base Presentar")
    ws.sheet_properties.tabColor = "E2EFDA"

    cols = df_pres.columns
    write_headers(ws, 1, cols)

    for r_idx, row_data in enumerate(df_pres.iter_rows(), start=2):
        fill = FILL_LIGHT if r_idx % 2 == 0 else FILL_WHITE
        fills = [fill] * len(cols)
        fmts = []
        for i, col in enumerate(cols):
            if col in ("ingreso_estimado_2025", "ticket_promedio_2025", "valor_balance_2024"):
                fmts.append("$#,##0.00")
            elif col == "numero_ruc":
                fmts.append("0")
            else:
                fmts.append(None)
        write_row(ws, r_idx, list(row_data), fills=fills, fmt=fmts)

    # Ajustar anchos
    col_widths_map = {
        "numero_ruc": 16, "razon_social": 30, "nombre_comercial": 30,
        "clase_contribuyente": 16, "tipo_contribuyente": 18,
        "actividad_economica": 50, "valor_balance_2024": 20,
        "ticket_promedio_2025": 20, "ingreso_estimado_2025": 22,
        "tipo_actividad": 28, "descripcion_corta": 40,
    }
    for i, col in enumerate(cols):
        ws.column_dimensions[get_column_letter(i + 1)].width = col_widths_map.get(col, 16)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{df_pres.height + 1}"


# ── Main ─────────────────────────────────────────────────────────────
def generar_resumen():
    print("Leyendo base_presentar...")
    df_pres = pl.read_parquet(BASE_PRESENTAR)

    print(f"  base_presentar: {df_pres['numero_ruc'].n_unique():,} RUCs")

    wb = Workbook()

    print("Generando hoja Resumen...")
    hoja_resumen(wb, df_pres)

    print("Generando hoja Glosario...")
    hoja_glosario(wb)

    print("Generando hoja Base Presentar...")
    hoja_base_presentar(wb, df_pres)

    wb.save(RESUMEN_XLSX)
    print(f"Reporte guardado en: {RESUMEN_XLSX}")


if __name__ == "__main__":
    generar_resumen()
